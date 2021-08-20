# ---
# jupyter:
#   jupytext:
#     cell_metadata_filter: all,-execution,-papermill,-trusted
#     formats: ipynb,py//py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.7.1
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %% [markdown] tags=[]
# # Description

# %% [markdown] tags=[]
# This notebook generates more end-user-friendly Excel files of some of the data generated in PhenoPLIER, such as the LV-gene matrix and LV-pathways.

# %% [markdown] tags=[]
# # Modules loading

# %% tags=[]
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

import conf
from utils import get_git_repository_path

# %% [markdown] tags=[]
# # Settings

# %% tags=[]
DELIVERABLES_BASE_DIR = get_git_repository_path() / "data"
display(DELIVERABLES_BASE_DIR)

# %% tags=[]
OUTPUT_DIR = DELIVERABLES_BASE_DIR / "multiplier"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
display(OUTPUT_DIR)

# %% [markdown] tags=[]
# # Load data

# %% [markdown] tags=[]
# ## MultiPLIER summary

# %% tags=[]
multiplier_model_summary = pd.read_pickle(conf.MULTIPLIER["MODEL_SUMMARY_FILE"])

# %% tags=[]
multiplier_model_summary.shape

# %% tags=[]
multiplier_model_summary.head()

# %% [markdown] tags=[]
# ## Get pathway-aligned LVs

# %% tags=[]
well_aligned_lvs = multiplier_model_summary[multiplier_model_summary["FDR"] < 0.05]

display(well_aligned_lvs.shape)
display(well_aligned_lvs.head())

# %% tags=[]
well_aligned_lv_codes = set([f"LV{lvi}" for lvi in well_aligned_lvs["LV index"]])

# %% tags=[]
len(well_aligned_lv_codes)

# %% tags=[]
list(well_aligned_lv_codes)[:5]

# %% [markdown] tags=[]
# ## MultiPLIER Z (gene loadings)

# %% tags=[]
multiplier_z = pd.read_pickle(conf.MULTIPLIER["MODEL_Z_MATRIX_FILE"])

# %% tags=[]
multiplier_z.shape

# %% tags=[]
multiplier_z.head()

# %% [markdown] tags=[]
# # Create LV-pathway dataframe

# %% tags=[]
lv_pathway_df = multiplier_model_summary[
    ["LV index", "pathway", "AUC", "p-value", "FDR"]
]

# %% tags=[]
lv_pathway_df["LV index"] = lv_pathway_df["LV index"].astype(int)

# %% tags=[]
lv_pathway_df = lv_pathway_df.sort_values(["LV index", "FDR"])

# %% tags=[]
lv_pathway_df["LV index"] = lv_pathway_df["LV index"].apply(lambda x: f"LV{x}")

# %% tags=[]
lv_pathway_df.shape

# %% tags=[]
lv_pathway_df = lv_pathway_df.rename(
    columns={"LV index": "LV identifier", "pathway": "Pathway"}
)

# %% tags=[]
lv_pathway_df.head()

# %% tags=[]
lv_pathway_df.tail()

# %% [markdown] tags=[]
# ## Save

# %% tags=[]
output_file = OUTPUT_DIR / "lv-pathways.xlsx"
display(output_file)

# %% tags=[]
lv_pathway_df.to_excel(output_file, index=False)

# %% tags=[]
# adjust column widths
wb = openpyxl.load_workbook(filename=output_file)
worksheet = wb.active

for col in worksheet.columns:
    max_length = 0
    column = get_column_letter(col[0].column)  # Get the column name
    for cell in col:
        if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
            continue

        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.05
    worksheet.column_dimensions[column].width = adjusted_width

wb.save(output_file)

# %% [markdown] tags=[]
# # Create LV-gene dataframe

# %% tags=[]
df = (
    multiplier_z.unstack()
    .to_frame()
    .reset_index()
    .rename(columns={0: "Weight", "level_0": "LV identifier", "level_1": "Gene symbol"})
)

# %% tags=[]
df = df.assign(lv_index=df["LV identifier"].apply(lambda x: int(x[2:])))

# %% tags=[]
df = df.sort_values(["lv_index", "Weight"], ascending=[True, False]).drop(
    columns=["lv_index"]
)

# %% tags=[]
df.shape

# %% tags=[]
df.head()

# %% tags=[]
df.tail()

# %% [markdown] tags=[]
# ## Save as TSV

# %% tags=[]
# output_file = OUTPUT_DIR / "lv-gene_weights.tsv.gz"
# display(output_file)

# %% tags=[]
# df.to_csv(output_file, sep="\t", index=False)

# %% [markdown] tags=[]
# ## Save as Excel

# %% tags=[]
output_file = OUTPUT_DIR / "lv-gene_weights.xlsx"
display(output_file)

# %% tags=[]
with pd.ExcelWriter(output_file) as writer:
    for lv_id in df["LV identifier"].unique():
        print(lv_id, end=", ", flush=True)

        lv_data = df[df["LV identifier"] == lv_id].drop(columns=["LV identifier"])
        lv_data = lv_data.sort_values("Weight", ascending=False)

        lv_data.to_excel(writer, index=False, sheet_name=lv_id)

# %% tags=[]
# adjust column widths
wb = openpyxl.load_workbook(filename=output_file)

for worksheet in wb.worksheets:
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                continue

            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.05
        worksheet.column_dimensions[column].width = adjusted_width

wb.save(output_file)

# %% tags=[]
